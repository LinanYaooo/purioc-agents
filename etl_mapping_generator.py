"""
ETL Mapping Excel Generator
============================
使用 openpyxl 生成包含复杂数据验证和条件格式的ETL Mapping Excel模板
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
import json


class ETLMappingExcelGenerator:
    """ETL Mapping Excel模板生成器"""
    
    def __init__(self):
        self.wb = Workbook()
        # 定义样式
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    def _add_data_validation(self, ws, cell_range, validation_type, **kwargs):
        """添加数据验证"""
        if validation_type == "list":
            formula = kwargs.get('formula', '')
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = kwargs.get('error', '输入值不在允许范围内')
            dv.errorTitle = kwargs.get('error_title', '无效输入')
            dv.prompt = kwargs.get('prompt', '')
            dv.promptTitle = kwargs.get('prompt_title', '')
            ws.add_data_validation(dv)
            dv.add(cell_range)
            
    def _add_conditional_formatting(self, ws, cell_range, formula, fill):
        """添加条件格式"""
        rule = FormulaRule(formula=[formula], fill=fill)
        ws.conditional_formatting.add(cell_range, rule)
        
    def create_entity_mapping_sheet(self):
        """创建实体级Mapping Sheet"""
        ws = self.wb.active
        ws.title = "实体级Mapping"
        
        headers = [
            "实体ID", "源系统", "源实体名", "源实体类型", 
            "目标系统", "目标实体名", "目标实体类型",
            "Mapping类型", "状态", "负责人", "备注"
        ]
        
        col_widths = [12, 15, 25, 15, 15, 25, 15, 15, 12, 15, 30]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
            
        # 数据验证
        systems = '"Oracle,MySQL,SQL Server,PostgreSQL,MongoDB,REST API,File,S3"'
        self._add_data_validation(ws, 'B2:B1000', 'list', formula=systems)
        self._add_data_validation(ws, 'E2:E1000', 'list', formula=systems)
        
        entity_types = '"Table,View,API Endpoint,File,Queue,Topic"'
        self._add_data_validation(ws, 'D2:D1000', 'list', formula=entity_types)
        self._add_data_validation(ws, 'G2:G1000', 'list', formula=entity_types)
        
        mapping_types = '"Direct,Transform,Join,Union,Aggregate,Lookup"'
        self._add_data_validation(ws, 'H2:H1000', 'list', formula=mapping_types)
        
        statuses = '"Draft,In Progress,Review,Approved,Deprecated"'
        self._add_data_validation(ws, 'I2:I1000', 'list', formula=statuses)
        
        # 条件格式
        self._add_conditional_formatting(
            ws, 'A2:A1000', 'ISBLANK(A2)',
            PatternFill(start_color="FFC7CE", fill_type="solid")
        )
        
        tab = Table(displayName="EntityMapping", ref="A1:K1000")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        return ws
        
    def create_attribute_mapping_sheet(self):
        """创建属性级Mapping Sheet"""
        ws = self.wb.create_sheet("属性级Mapping")
        
        headers = [
            "MappingID", "实体ID", 
            "源字段路径", "源字段名", "源数据类型", "源长度", "源精度", "源必填",
            "目标字段路径", "目标字段名", "目标数据类型", "目标长度", "目标精度", "目标必填",
            "转换规则", "默认值", "业务规则", "数据质量规则", "状态", "备注"
        ]
        
        col_widths = [12, 12, 30, 20, 12, 10, 10, 10, 30, 20, 12, 10, 10, 10, 20, 15, 25, 25, 12, 30]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
            
        # 数据验证
        data_types = '"String,Integer,Decimal,Date,DateTime,Boolean,JSON,Array,Binary"'
        self._add_data_validation(ws, 'E2:E1000', 'list', formula=data_types)
        self._add_data_validation(ws, 'K2:K1000', 'list', formula=data_types)
        
        yes_no = '"Yes,No"'
        self._add_data_validation(ws, 'H2:H1000', 'list', formula=yes_no)
        self._add_data_validation(ws, 'N2:N1000', 'list', formula=yes_no)
        
        transform_rules = '"Direct Map,UpperCase,LowerCase,Trim,Substring,Date Format,Lookup,Calculate,Concat,Split,Custom"'
        self._add_data_validation(ws, 'O2:O1000', 'list', formula=transform_rules)
        
        statuses = '"Draft,In Progress,Review,Approved,Deprecated"'
        self._add_data_validation(ws, 'S2:S1000', 'list', formula=statuses)
        
        # 条件格式
        self._add_conditional_formatting(
            ws, 'A2:T1000', 'OR(ISBLANK(B2),ISBLANK(D2),ISBLANK(J2))',
            PatternFill(start_color="FFC7CE", fill_type="solid")
        )
        
        tab = Table(displayName="AttributeMapping", ref="A1:T1000")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        return ws
        
    def save(self, filename=None):
        """保存Excel文件"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ETL_Mapping_Template_{timestamp}.xlsx"
        self.wb.save(filename)
        print(f"Excel文件已生成: {filename}")
        return filename


def flatten_json(json_obj, parent_key='', sep='.'):
    """JSON扁平化"""
    items = []
    if isinstance(json_obj, dict):
        for key, value in json_obj.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                items.extend(flatten_json(value, new_key, sep))
            else:
                items.append({'path': new_key, 'name': key, 'type': type(value).__name__})
    return items


if __name__ == "__main__":
    generator = ETLMappingExcelGenerator()
    generator.create_entity_mapping_sheet()
    generator.create_attribute_mapping_sheet()
    generator.save()
    
    # JSON扁平化示例
    sample = {"customer": {"id": 123, "name": "John", "address": {"city": "Shanghai"}}}
    print("\nJSON扁平化示例:")
    for field in flatten_json(sample):
        print(f"  路径: {field['path']}, 名称: {field['name']}, 类型: {field['type']}")
