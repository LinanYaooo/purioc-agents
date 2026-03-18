"""
生成实体级Mapping Excel文件
根据用户提供的列名创建Excel
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "实体级Mapping"

# 定义样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定义列名（按用户提供的顺序）
headers = [
    "目标实体 所属schema*",
    "目标实体 物理表名称*",
    "目标实体 中文名称",
    "来源实体 所属schema*",
    "来源实体 中文名称",
    "来源实体 物理表名称*",
    "来源实体 别名",
    "关联类型*",
    "关联条件*",
    "具体关联",
    "责任人",
    "备注",
    "表名长度(需要小于40位)"
]

# 创建表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    
    # 标记必填项（带*的）
    if '*' in header:
        cell.fill = required_fill
        cell.font = Font(bold=True, color="FF0000", size=11)

# 添加数据验证
# 关联类型下拉列表
dv_join_type = DataValidation(
    type="list",
    formula1='"INNER,LEFT,RIGHT,FULL,CROSS"',
    allow_blank=True
)
dv_join_type.error = '请从下拉列表中选择关联类型'
dv_join_type.errorTitle = '输入错误'
ws.add_data_validation(dv_join_type)
dv_join_type.add('H2:H1000')  # 关联类型列

# 设置列宽
column_widths = {
    1: 20,   # 目标schema
    2: 30,   # 目标表名
    3: 25,   # 目标中文名
    4: 20,   # 来源schema
    5: 25,   # 来源中文名
    6: 30,   # 来源表名
    7: 15,   # 来源别名
    8: 12,   # 关联类型
    9: 40,   # 关联条件
    10: 35,  # 具体关联
    11: 12,  # 责任人
    12: 30,  # 备注
    13: 20   # 表名长度
}

for col_num, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

# 设置行高
ws.row_dimensions[1].height = 35

# 添加示例数据行
example_data = [
    "dws",                    # 目标schema
    "dws_trade_orders_summary",  # 目标表名
    "交易订单汇总表",          # 目标中文名
    "dwd",                    # 来源schema
    "交易订单明细表",          # 来源中文名
    "dwd_trade_orders_detail",  # 来源表名
    "a",                      # 来源别名
    "LEFT",                   # 关联类型
    "a.channel_code = b.channel_code",  # 关联条件
    "LEFT JOIN dim_channel_info b ON a.channel_code = b.channel_code",  # 具体关联
    "张三",                   # 责任人
    "每日汇总",               # 备注
    "26"                      # 表名长度
]

for col, value in enumerate(example_data, 1):
    cell = ws.cell(row=2, column=col, value=value)
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 冻结首行
ws.freeze_panes = 'A2'

# 保存文件
output_path = "实体级Mapping.xlsx"
wb.save(output_path)

print(f"[OK] Excel文件已生成: {output_path}")
print(f"\n[文件信息]")
print(f"   - Sheet名称: 实体级Mapping")
print(f"   - 列数: {len(headers)}")
print(f"   - 包含数据验证: 关联类型下拉列表")
print(f"   - 包含示例数据: 1行")
print(f"\n[列名列表]")
for i, header in enumerate(headers, 1):
    required = "(必填)" if '*' in header else ""
    print(f"   {i}. {header} {required}")
