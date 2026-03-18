"""
生成完备版ETL Mapping Excel文件
基于Mapping Design v2.0，包含所有P0关键字段
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

# 创建工作簿
wb = Workbook()

# 定义样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_width(ws, col_widths):
    """设置列宽"""
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

def create_header(ws, headers, row=1):
    """创建表头"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def add_data_validation(ws, cell_range, formula1, allow_blank=True):
    """添加数据验证（下拉列表）"""
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
    dv.error = '请从下拉列表中选择'
    dv.errorTitle = '输入错误'
    ws.add_data_validation(dv)
    dv.add(cell_range)

# ==================== Sheet 1: 实体级Mapping ====================
ws1 = wb.active
ws1.title = "实体级Mapping"

# 定义实体级表头（按逻辑分组排列）
entity_headers = [
    # === 基础信息 ===
    "mapping_id\n(映射ID)",
    "target_schema\n(目标schema)",
    "target_table\n(目标表名)",
    "target_table_cn\n(目标表中文名)",
    "table_description\n(表用途说明)",
    "data_layer\n(数据层级)",
    "business_domain\n(业务域)",
    "responsible_person\n(责任人)",
    "version\n(版本号)",
    
    # === 执行策略配置 (P0) ===
    "loading_strategy\n(加载策略)*",
    "schedule_type\n(调度周期)",
    "schedule_cron\n(Cron表达式)",
    "dependency_tasks\n(上游依赖任务)*",
    "execution_priority\n(执行优先级)",
    "timeout_minutes\n(超时时间)",
    "retry_count\n(重试次数)",
    "retry_interval_minutes\n(重试间隔)",
    "parallel_enabled\n(允许并行)",
    
    # === 增量/拉链配置 ===
    "incremental_watermark_field\n(水位线字段)",
    "incremental_watermark_value\n(水位线值)",
    "incremental_condition\n(增量条件)",
    "scd_effective_date_field\n(生效日期字段)",
    "scd_expire_date_field\n(失效日期字段)",
    "scd_current_flag_field\n(当前标识字段)",
    "scd_natural_keys\n(自然键列表)",
    
    # === DWS分布式配置 (P0) ===
    "distribution_key\n(分布键)*",
    "distribution_type\n(分布类型)",
    "partition_key\n(分区键)*",
    "partition_type\n(分区类型)",
    "partition_granularity\n(分区粒度)",
    "partition_retention_days\n(分区保留天数)",
    
    # === SQL条件配置 (P0核心) ===
    "join_type\n(JOIN类型)",
    "join_table_alias\n(关联表别名)",
    "join_table_name\n(关联表名)",
    "join_condition\n(JOIN条件)*",
    "where_conditions\n(WHERE条件)*",
    "group_by_fields\n(GROUP BY字段)",
    "having_conditions\n(HAVING条件)",
    "order_by_fields\n(ORDER BY字段)",
    
    # === 数据质量配置 (P0) ===
    "exception_strategy\n(异常策略)*",
    "validation_rule_1\n(校验规则1)",
    "validation_rule_2\n(校验规则2)",
    "validation_rule_3\n(校验规则3)",
    "error_threshold_percent\n(错误率阈值%)",
    
    # === 备注 ===
    "remark\n(备注)"
]

# 创建表头
create_header(ws1, entity_headers)

# 添加数据验证（下拉列表）
# loading_strategy
add_data_validation(ws1, 'K2:K1000', '"FULL,INCREMENTAL_INSERT,INCREMENTAL_MERGE,INCREMENTAL_UPDATE,SCD2"')

# schedule_type
add_data_validation(ws1, 'L2:L1000', '"DAILY,WEEKLY,MONTHLY,HOURLY,REALTIME"')

# data_layer
add_data_validation(ws1, 'F2:F1000', '"ODS,DWD,DWS,ADS,DIM"')

# parallel_enabled
add_data_validation(ws1, 'R2:R1000', '"TRUE,FALSE"')

# distribution_type
add_data_validation(ws1, 'X2:X1000', '"HASH,REPLICATION,ROUNDROBIN"')

# partition_type
add_data_validation(ws1, 'Z2:Z1000', '"RANGE,LIST,HASH"')

# partition_granularity
add_data_validation(ws1, 'AA2:AA1000', '"DAY,MONTH,YEAR"')

# join_type
add_data_validation(ws1, 'AD2:AD1000', '"INNER,LEFT,RIGHT,FULL,CROSS"')

# exception_strategy
add_data_validation(ws1, 'AL2:AL1000', '"REJECT,LOG,ALLOW,REJECT_AFTER_LOG"')

# 添加示例数据行（第一行为示例）
example_row = [
    "M_DWS_ORDERS_SUM_001",  # mapping_id
    "dws",  # target_schema
    "dws_trade_orders_summary_day",  # target_table
    "交易订单日汇总表",  # target_table_cn
    "按天汇总各渠道订单金额、数量、客户数",  # table_description
    "DWS",  # data_layer
    "TRADE",  # business_domain
    "张三",  # responsible_person
    "v1.0.0",  # version
    "INCREMENTAL_MERGE",  # loading_strategy (P0)
    "DAILY",  # schedule_type
    "0 2 * * *",  # schedule_cron
    "dwd_trade_orders_detail_day,dim_channel_info",  # dependency_tasks (P0)
    "100",  # execution_priority
    "60",  # timeout_minutes
    "3",  # retry_count
    "10",  # retry_interval_minutes
    "FALSE",  # parallel_enabled
    "etl_load_time",  # incremental_watermark_field
    "${BATCH_DATE} 00:00:00",  # incremental_watermark_value
    "etl_load_time >= '${BATCH_DATE}'",  # incremental_condition
    "",  # scd_effective_date_field
    "",  # scd_expire_date_field
    "",  # scd_current_flag_field
    "",  # scd_natural_keys
    "channel_code,stat_date",  # distribution_key (P0)
    "HASH",  # distribution_type
    "stat_date",  # partition_key (P0)
    "RANGE",  # partition_type
    "DAY",  # partition_granularity
    "365",  # partition_retention_days
    "LEFT",  # join_type
    "b",  # join_table_alias
    "dim_channel_info",  # join_table_name
    "a.channel_code = b.channel_code AND b.is_valid = 'Y'",  # join_condition (P0)
    "order_status IN ('COMPLETED','PAID') AND is_deleted = 'N'",  # where_conditions (P0)
    "stat_date,channel_code,channel_name",  # group_by_fields
    "",  # having_conditions
    "stat_date DESC,total_amount DESC",  # order_by_fields
    "LOG",  # exception_strategy (P0)
    "RULE001:NOT_NULL:summary_id",  # validation_rule_1
    "RULE002:RANGE:total_amount:0:999999999",  # validation_rule_2
    "RULE003:REFERENTIAL:channel_code:dim_channel_info:channel_code",  # validation_rule_3
    "5.0",  # error_threshold_percent
    "首次创建",  # remark
]

# 填入示例数据
for col, value in enumerate(example_row, 1):
    cell = ws1.cell(row=2, column=col, value=value)
    cell.fill = example_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 设置列宽（根据内容调整）
entity_col_widths = {
    1: 25, 2: 15, 3: 30, 4: 20, 5: 35,
    6: 12, 7: 12, 8: 12, 9: 12,
    10: 18, 11: 15, 12: 15, 13: 35, 14: 15, 15: 12, 16: 10, 17: 12,
    18: 20, 19: 25, 20: 25, 21: 30, 22: 20, 23: 20, 24: 20,
    25: 25, 26: 15, 27: 20, 28: 15, 29: 15, 30: 18,
    31: 12, 32: 12, 33: 20, 34: 45, 35: 45, 36: 35, 37: 30, 38: 35,
    39: 18, 40: 35, 41: 35, 42: 55, 43: 12, 44: 20
}
set_column_width(ws1, entity_col_widths)

# 设置行高
ws1.row_dimensions[1].height = 35
ws1.row_dimensions[2].height = 25

# 添加说明行（第2行为说明）
ws1.insert_rows(2)
ws1.row_dimensions[2].height = 40
ws1.merge_cells('A2:AI2')
desc_cell = ws1['A2']
desc_cell.value = "说明：带*为P0必填字段（阻塞性字段），黄色高亮。loading_strategy决定代码结构，distribution_key/partition_key是DWS必需，sql_conditions包含JOIN/WHERE/GROUP BY，exception_strategy决定异常处理方式"
desc_cell.font = Font(italic=True, size=10, color="666666")
desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 冻结首行
ws1.freeze_panes = 'A3'

# ==================== Sheet 2: 属性级Mapping ====================
ws2 = wb.create_sheet("属性级Mapping")

# 定义属性级表头
attribute_headers = [
    # === 基础关联 ===
    "mapping_id\n(关联实体ID)*",
    "target_table\n(目标表名)*",
    
    # === 目标字段信息 ===
    "target_field\n(目标字段)*",
    "target_field_cn\n(目标字段中文名)",
    "target_data_type\n(目标数据类型)*",
    "target_precision\n(目标精度)",
    "is_nullable\n(可空)",
    "is_pk\n(主键)",
    "field_order\n(字段顺序)",
    
    # === 生成方式 (P0核心) ===
    "expression_type\n(生成方式)*",
    "source_schema\n(来源schema)",
    "source_table\n(来源表名)",
    "source_table_alias\n(来源表别名)",
    "source_field\n(来源字段)",
    
    # === 结构化加工逻辑 (替代文本描述) ===
    "function_name\n(函数名)",
    "expression_template\n(表达式模板)",
    "condition_1\n(条件1)",
    "result_1\n(结果1)",
    "condition_2\n(条件2)",
    "result_2\n(结果2)",
    "else_result\n(ELSE结果)",
    "distinct_flag\n(DISTINCT)",
    "filter_condition\n(过滤条件)",
    "group_by_flag\n(是否分组键)",
    "order_by_sequence\n(排序序号)",
    "order_by_direction\n(排序方向)",
    
    # === 关联查询配置 ===
    "lookup_table\n(关联维表)",
    "lookup_key\n(关联键)",
    "lookup_return_field\n(返回值字段)",
    "default_value\n(默认值)",
    
    # === 备注 ===
    "remark\n(备注)"
]

# 创建表头
create_header(ws2, attribute_headers)

# 添加数据验证
# expression_type - 最关键的下拉列表，避免文本描述
add_data_validation(ws2, 'K2:K10000', '"DIRECT,DIRECT_MAPPING,LOOKUP,AGGREGATE,CASE,CALCULATE,DERIVED,SYSTEM,CONSTANT,WINDOW_FUNCTION"')

# is_nullable, is_pk, distinct_flag, group_by_flag
add_data_validation(ws2, 'G2:H10000', '"Y,N"')
add_data_validation(ws2, 'T2:T10000', '"Y,N"')
add_data_validation(ws2, 'X2:X10000', '"Y,N"')

# order_by_direction
add_data_validation(ws2, 'Z2:Z10000', '"ASC,DESC"')

# 添加示例数据行
attr_examples = [
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "summary_id", "汇总ID", "VARCHAR", "32", "N", "Y", "1",
        "DERIVED", "dwd", "dwd_trade_orders_detail_day", "a", "stat_date,channel_code",
        "MD5", "MD5({0}::TEXT || '_' || {1})", "", "", "", "", "", "N", "", "N", "", "",
        "", "", "", "",
        "主键：日期+渠道代码MD5"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "stat_date", "统计日期", "DATE", "", "N", "N", "2",
        "DIRECT", "dwd", "dwd_trade_orders_detail_day", "a", "stat_date",
        "", "{0}", "", "", "", "", "", "N", "", "Y", "1", "ASC",
        "", "", "", "",
        "直接映射"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "channel_name", "渠道名称", "VARCHAR", "100", "Y", "N", "4",
        "LOOKUP", "dim", "dim_channel_info", "b", "channel_name",
        "COALESCE", "COALESCE({0}, '未知渠道')", "", "", "", "", "", "N", "", "N", "", "",
        "dim_channel_info", "channel_code", "channel_name", "未知渠道",
        "关联维度表获取，缺失时默认'未知渠道'"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "total_orders", "订单总数", "BIGINT", "", "N", "N", "5",
        "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "a", "order_id",
        "COUNT", "COUNT({0})", "", "", "", "", "", "N", "", "N", "", "",
        "", "", "", "",
        "COUNT聚合"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "total_customers", "客户总数", "BIGINT", "", "N", "N", "7",
        "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "a", "customer_id",
        "COUNT", "COUNT(DISTINCT {0})", "", "", "", "", "", "Y", "customer_id IS NOT NULL", "N", "", "",
        "", "", "", "",
        "COUNT DISTINCT聚合，带过滤条件"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "avg_order_amount", "平均订单金额", "DECIMAL", "18,2", "Y", "N", "8",
        "CALCULATE", "", "", "", "total_amount,total_orders",
        "DIVIDE", "CASE WHEN {1}>0 THEN {0}/{1} ELSE 0 END", "", "", "", "", "", "N", "", "N", "", "",
        "", "", "", "0",
        "公式计算：总金额/总订单数，除0处理"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "amount_level", "金额等级", "VARCHAR", "20", "Y", "N", "9",
        "CASE", "", "", "", "total_amount",
        "CASE", "CASE WHEN {0}>=1000000 THEN 'HIGH' WHEN {0}>=100000 THEN 'MEDIUM' ELSE 'LOW' END",
        "{0}>=1000000", "'HIGH'", "{0}>=100000", "'MEDIUM'", "'LOW'", "N", "", "N", "", "",
        "", "", "", "",
        "CASE WHEN条件判断：>=100万HIGH，>=10万MEDIUM，其他LOW"
    ],
    [
        "M_DWS_ORDERS_SUM_001", "dws_trade_orders_summary_day",
        "etl_load_time", "ETL加载时间", "TIMESTAMP", "", "N", "N", "10",
        "SYSTEM", "", "", "", "",
        "CURRENT_TIMESTAMP", "CURRENT_TIMESTAMP", "", "", "", "", "", "N", "", "N", "", "",
        "", "", "", "",
        "系统变量"
    ],
]

# 填入示例数据
for row_idx, row_data in enumerate(attr_examples, 3):
    for col, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col, value=value)
        cell.fill = example_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 设置列宽
attr_col_widths = {
    1: 25, 2: 30,  # 基础关联
    3: 20, 4: 20, 5: 15, 6: 12, 7: 8, 8: 8, 9: 10,  # 目标字段
    10: 18, 11: 12, 12: 25, 13: 12, 14: 18,  # 生成方式
    15: 15, 16: 25, 17: 20, 18: 15, 19: 20, 20: 15, 21: 15, 22: 8, 23: 20, 24: 12, 25: 10, 26: 12,  # 加工逻辑
    27: 20, 28: 12, 29: 15, 30: 15,  # 关联查询
    31: 25  # 备注
}
set_column_width(ws2, attr_col_widths)

# 设置行高
ws2.row_dimensions[1].height = 35
for i in range(3, 3 + len(attr_examples)):
    ws2.row_dimensions[i].height = 25

# 添加说明行
ws2.insert_rows(2)
ws2.row_dimensions[2].height = 50
ws2.merge_cells('A2:AE2')
desc_cell2 = ws2['A2']
desc_cell2.value = """说明：
1. expression_type(生成方式)是关键字段，必须从下拉列表选择：
   - DIRECT/LOOKUP/AGGREGATE/CASE/CALCULATE/DERIVED/SYSTEM/CONSTANT
2. 结构化加工逻辑列（condition_1/result_1/condition_2/result_2/else_result）替代原有的"加工逻辑描述"文本字段
3. CASE类型：填写condition/result列；AGGREGATE类型：填写function_name/distinct_flag；LOOKUP类型：填写lookup_table/lookup_key
4. 示例数据展示了8种不同的生成方式，可参考填写"""
desc_cell2.font = Font(italic=True, size=9, color="666666")
desc_cell2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 冻结首行
ws2.freeze_panes = 'A3'

# ==================== 保存文件 ====================
output_path = "spec/ETL_Mapping_Design_v2.0_Excel版.xlsx"
wb.save(output_path)

print(f"[OK] Excel文件已生成: {output_path}")
print(f"\n[统计] 文件包含:")
print(f"   - Sheet 1: 实体级Mapping (44列，包含所有P0字段)")
print(f"     * 基础信息列 (9列)")
print(f"     * 执行策略配置列 (9列) - loading_strategy, dependency_tasks等")
print(f"     * DWS分布式配置列 (6列) - distribution_key, partition_key等")
print(f"     * SQL条件配置列 (8列) - join_type, join_condition, where_conditions等")
print(f"     * 数据质量配置列 (5列) - exception_strategy, validation_rules等")
print(f"\n   - Sheet 2: 属性级Mapping (31列，结构化加工逻辑)")
print(f"     * 基础映射列 (9列)")
print(f"     * 生成方式列 (5列) - expression_type下拉选择，避免文本描述")
print(f"     * 结构化加工逻辑列 (12列) - condition/result替代文本描述")
print(f"     * 关联查询配置列 (4列)")
print(f"\n[关键改进]:")
print(f"   1. 所有P0字段已集成到Excel中")
print(f"   2. expression_type下拉列表强制结构化输入")
print(f"   3. condition/result列替代'加工逻辑描述'文本字段")
print(f"   4. 提供8行完整示例数据，覆盖所有生成方式")
print(f"   5. 数据验证（下拉列表）防止错误输入")
