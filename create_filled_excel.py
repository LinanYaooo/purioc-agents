#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
填充Mapping Excel文档
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

def create_entity_mapping_excel():
    """创建填充好的实体级Mapping Excel"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "实体级Mapping"
    
    # 设置列宽
    column_widths = {
        'A': 8,   # 序号
        'B': 35,  # 目标实体 物理表名称
        'C': 30,  # 目标实体 中文名称
        'D': 15,  # Schema
        'E': 15,  # 加载策略
        'F': 15,  # 调度周期
        'G': 40,  # 上游依赖任务
        'H': 25,  # 分布键
        'I': 15,  # 分区键
        'J': 15,  # 异常策略
        'K': 50,  # 来源表配置(JSON)
        'L': 80,  # SQL条件配置(JSON)
        'M': 15,  # 责任人
        'N': 15,  # 版本号
        'O': 20,  # 创建时间
        'P': 30,  # 备注
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 写入表头
    headers = [
        "序号", "目标实体 物理表名称*", "目标实体 中文名称", "目标Schema*",
        "加载策略*", "调度周期*", "上游依赖任务*", "分布键*", 
        "分区键*", "异常处理策略*", "来源表配置(JSON)*", "SQL条件配置(JSON)*",
        "责任人*", "版本号*", "创建时间", "备注"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 准备数据
    entity_data = {
        "target_table": "dws_ecommerce_channel_sales_analysis_day",
        "target_table_cn": "电商渠道销售分析日汇总表",
        "schema": "dws",
        "loading_strategy": "INCREMENTAL_MERGE",
        "schedule_type": "DAILY",
        "dependency_tasks": "dwd_trade_orders_detail_day, dwd_order_items_detail_day, dwd_payments_detail_day, dwd_refunds_detail_day, dwd_inventory_detail_day, dim_channel_info, dim_product_info, dim_customer_info, dim_region_info, dim_promotion_info, dim_store_info, dim_category_info",
        "distribution_key": "analysis_id",
        "partition_key": "stat_date",
        "exception_strategy": "LOG",
        "source_tables": json.dumps([
            {"source_id": "S001", "schema": "dwd", "table_name": "dwd_trade_orders_detail_day", "alias": "ord", "table_type": "PRIMARY", "description": "订单明细事实表"},
            {"source_id": "S002", "schema": "dwd", "table_name": "dwd_order_items_detail_day", "alias": "itm", "table_type": "DETAIL", "description": "订单商品明细表"},
            {"source_id": "S003", "schema": "dwd", "table_name": "dwd_payments_detail_day", "alias": "pay", "table_type": "DETAIL", "description": "支付明细表"},
            {"source_id": "S004", "schema": "dwd", "table_name": "dwd_refunds_detail_day", "alias": "ref", "table_type": "DETAIL", "description": "退款明细表"},
            {"source_id": "S005", "schema": "dwd", "table_name": "dwd_inventory_detail_day", "alias": "inv", "table_type": "DETAIL", "description": "库存明细表"},
            {"source_id": "S006", "schema": "dim", "table_name": "dim_channel_info", "alias": "chn", "table_type": "LOOKUP", "description": "渠道维度表"},
            {"source_id": "S007", "schema": "dim", "table_name": "dim_product_info", "alias": "prd", "table_type": "LOOKUP", "description": "商品维度表"},
            {"source_id": "S008", "schema": "dim", "table_name": "dim_customer_info", "alias": "cst", "table_type": "LOOKUP", "description": "客户维度表"},
            {"source_id": "S009", "schema": "dim", "table_name": "dim_region_info", "alias": "reg", "table_type": "LOOKUP", "description": "地区维度表"},
            {"source_id": "S010", "schema": "dim", "table_name": "dim_promotion_info", "alias": "prm", "table_type": "LOOKUP", "description": "促销维度表"},
            {"source_id": "S011", "schema": "dim", "table_name": "dim_store_info", "alias": "sto", "table_type": "LOOKUP", "description": "门店维度表"},
            {"source_id": "S012", "schema": "dim", "table_name": "dim_category_info", "alias": "cat", "table_type": "LOOKUP", "description": "类目维度表"}
        ], ensure_ascii=False),
        "sql_conditions": json.dumps({
            "joins": [
                {"join_sequence": 1, "join_type": "LEFT", "table_alias": "itm", "join_condition": "ord.order_id = itm.order_id AND itm.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 2, "join_type": "LEFT", "table_alias": "pay", "join_condition": "ord.order_id = pay.order_id AND pay.pay_status = 'SUCCESS'", "join_hint": "USE_HASH"},
                {"join_sequence": 3, "join_type": "LEFT", "table_alias": "ref", "join_condition": "ord.order_id = ref.order_id AND ref.refund_status = 'COMPLETED'", "join_hint": "USE_HASH"},
                {"join_sequence": 4, "join_type": "LEFT", "table_alias": "inv", "join_condition": "itm.product_id = inv.product_id AND inv.stat_date = ord.stat_date", "join_hint": "USE_HASH"},
                {"join_sequence": 5, "join_type": "LEFT", "table_alias": "chn", "join_condition": "ord.channel_code = chn.channel_code AND chn.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 6, "join_type": "LEFT", "table_alias": "prd", "join_condition": "itm.product_id = prd.product_id AND prd.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 7, "join_type": "LEFT", "table_alias": "cst", "join_condition": "ord.customer_id = cst.customer_id", "join_hint": "USE_HASH"},
                {"join_sequence": 8, "join_type": "LEFT", "table_alias": "reg", "join_condition": "ord.region_code = reg.region_code AND reg.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 9, "join_type": "LEFT", "table_alias": "prm", "join_condition": "ord.promotion_id = prm.promotion_id AND prm.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 10, "join_type": "LEFT", "table_alias": "sto", "join_condition": "ord.store_code = sto.store_code AND sto.is_valid = 'Y'", "join_hint": "USE_HASH"},
                {"join_sequence": 11, "join_type": "LEFT", "table_alias": "cat", "join_condition": "prd.category_id = cat.category_id AND cat.is_valid = 'Y'", "join_hint": "USE_HASH"}
            ],
            "where": [
                {"sequence": 1, "condition_group": 1, "logic": "AND", "field": "ord.order_status", "operator": "IN", "value": ["COMPLETED", "PAID", "SHIPPED"], "value_type": "ARRAY"},
                {"sequence": 2, "condition_group": 1, "logic": "AND", "field": "ord.is_deleted", "operator": "=", "value": "N", "value_type": "STRING"},
                {"sequence": 3, "condition_group": 1, "logic": "AND", "field": "ord.order_type", "operator": "!=", "value": "TEST", "value_type": "STRING"}
            ],
            "group_by": [
                {"sequence": 1, "expression": "ord.stat_date", "alias": None},
                {"sequence": 2, "expression": "ord.channel_code", "alias": None},
                {"sequence": 3, "expression": "chn.channel_name", "alias": None},
                {"sequence": 4, "expression": "chn.channel_type", "alias": None},
                {"sequence": 5, "expression": "reg.region_name", "alias": None},
                {"sequence": 6, "expression": "cat.category_name", "alias": None}
            ],
            "having": [
                {"sequence": 1, "logic": "AND", "condition": "COUNT(DISTINCT ord.order_id) > 0", "description": "只保留有订单记录的渠道"}
            ],
            "order_by": [
                {"sequence": 1, "expression": "ord.stat_date", "direction": "DESC"},
                {"sequence": 2, "expression": "total_net_amount", "direction": "DESC"}
            ]
        }, ensure_ascii=False),
        "owner": "数据开发工程师",
        "version": "v1.0.0",
        "create_time": "2026-03-18",
        "remark": "电商全渠道销售分析场景，12个来源表，复杂多表JOIN关联"
    }
    
    # 写入数据行
    row_data = [
        1,
        entity_data["target_table"],
        entity_data["target_table_cn"],
        entity_data["schema"],
        entity_data["loading_strategy"],
        entity_data["schedule_type"],
        entity_data["dependency_tasks"],
        entity_data["distribution_key"],
        entity_data["partition_key"],
        entity_data["exception_strategy"],
        entity_data["source_tables"],
        entity_data["sql_conditions"],
        entity_data["owner"],
        entity_data["version"],
        entity_data["create_time"],
        entity_data["remark"]
    ]
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 120
    
    # 保存
    wb.save("实体级Mapping_filled.xlsx")
    print("[OK] 实体级Mapping_filled.xlsx 已创建")

def create_attribute_mapping_excel():
    """创建填充好的属性级Mapping Excel"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "属性级Mapping"
    
    # 设置列宽
    column_widths = {
        'A': 6,   # 序号
        'B': 35,  # 目标实体 物理表名称
        'C': 30,  # 目标实体 字段名称
        'D': 25,  # 目标实体 属性名称
        'E': 12,  # 生成方式
        'F': 12,  # 来源schema
        'G': 35,  # 来源表名称
        'H': 10,  # 来源别名
        'I': 25,  # 来源字段名称
        'J': 25,  # 来源属性名称
        'K': 60,  # 加工逻辑描述
        'L': 10,  # 去重判断
        'M': 15,  # 责任人
        'N': 12,  # 字段类型
        'O': 12,  # 字段长度
        'P': 12,  # 主键
        'Q': 25,  # 变更记录
        'R': 30,  # 备注
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置表头样式
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 写入表头
    headers = [
        "序号", "目标实体 物理表名称*", "目标实体 字段名称*", "目标实体 属性名称",
        "生成方式*", "来源实体 所属schema*", "来源实体 物理表名称*", "来源实体 别名",
        "来源实体 字段名称*", "来源实体 属性名称", "加工逻辑描述*", "去重判断",
        "责任人", "字段类型", "字段长度", "主键(Y,N)", "变更记录", "备注"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 属性数据
    attributes = [
        (1, "analysis_id", "分析记录主键", "DERIVED", "", "", "", "", "", "MD5(stat_date || '_' || channel_code)", "", "STRING", "50", "Y", "2026-03-18 v1.0.0 初始创建", "主键：日期+渠道MD5"),
        (2, "stat_date", "统计日期", "DIRECT", "dwd", "dwd_trade_orders_detail_day", "ord", "stat_date", "", "直接映射：ord.stat_date", "", "DATE", "", "N", "2026-03-18 v1.0.0 初始创建", "统计日期"),
        (3, "channel_code", "渠道代码", "DIRECT", "dwd", "dwd_trade_orders_detail_day", "ord", "channel_code", "", "直接映射：ord.channel_code", "", "STRING", "20", "N", "2026-03-18 v1.0.0 初始创建", "渠道代码"),
        (4, "channel_name", "渠道名称", "LOOKUP", "dim", "dim_channel_info", "chn", "channel_name", "", "COALESCE(chn.channel_name, '未知渠道')", "", "STRING", "100", "N", "2026-03-18 v1.0.0 初始创建", "渠道名称"),
        (5, "channel_type", "渠道类型", "LOOKUP", "dim", "dim_channel_info", "chn", "channel_type", "", "COALESCE(chn.channel_type, 'OTHER')", "", "STRING", "20", "N", "2026-03-18 v1.0.0 初始创建", "渠道类型：APP/小程序/H5/线下"),
        (6, "region_name", "地区名称", "LOOKUP", "dim", "dim_region_info", "reg", "region_name", "", "COALESCE(reg.region_name, '未知地区')", "", "STRING", "100", "N", "2026-03-18 v1.0.0 初始创建", "地区名称"),
        (7, "category_name", "类目名称", "LOOKUP", "dim", "dim_category_info", "cat", "category_name", "", "COALESCE(cat.category_name, '未知类目')", "", "STRING", "100", "N", "2026-03-18 v1.0.0 初始创建", "类目名称"),
        (8, "total_orders", "订单总数", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "order_id", "", "COUNT(DISTINCT ord.order_id)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "订单总数（去重）"),
        (9, "total_order_amount", "订单总金额", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "order_amount", "", "SUM(ord.order_amount)", "", "DECIMAL", "18,2", "N", "2026-03-18 v1.0.0 初始创建", "订单总金额"),
        (10, "total_refund_amount", "退款总金额", "AGGREGATE", "dwd", "dwd_refunds_detail_day", "ref", "refund_amount", "", "COALESCE(SUM(ref.refund_amount), 0)", "", "DECIMAL", "18,2", "N", "2026-03-18 v1.0.0 初始创建", "退款总金额"),
        (11, "total_net_amount", "净销售额", "CALCULATE", "", "", "", "", "", "total_order_amount - total_refund_amount", "", "DECIMAL", "18,2", "N", "2026-03-18 v1.0.0 初始创建", "净销售额"),
        (12, "total_customers", "客户总数", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "customer_id", "", "COUNT(DISTINCT ord.customer_id)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "客户总数（去重）"),
        (13, "total_products", "商品总数", "AGGREGATE", "dwd", "dwd_order_items_detail_day", "itm", "product_id", "", "COUNT(DISTINCT itm.product_id)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "商品总数（去重）"),
        (14, "total_quantity", "商品总数量", "AGGREGATE", "dwd", "dwd_order_items_detail_day", "itm", "quantity", "", "SUM(itm.quantity)", "", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "商品总数量"),
        (15, "member_customers", "会员客户数", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "customer_id", "", "COUNT(DISTINCT CASE WHEN cst.customer_type = 'MEMBER' THEN ord.customer_id END)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "会员客户数"),
        (16, "non_member_customers", "非会员客户数", "CALCULATE", "", "", "", "", "", "total_customers - member_customers", "", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "非会员客户数"),
        (17, "avg_order_amount", "平均订单金额", "CALCULATE", "", "", "", "", "", "CASE WHEN total_orders > 0 THEN total_net_amount / total_orders ELSE 0 END", "", "DECIMAL", "18,2", "N", "2026-03-18 v1.0.0 初始创建", "平均订单金额"),
        (18, "avg_customer_value", "客户平均消费金额", "CALCULATE", "", "", "", "", "", "CASE WHEN total_customers > 0 THEN total_net_amount / total_customers ELSE 0 END", "", "DECIMAL", "18,2", "N", "2026-03-18 v1.0.0 初始创建", "客户平均消费金额"),
        (19, "channel_value_level", "渠道价值等级", "CASE", "", "", "", "", "", "CASE WHEN total_net_amount >= 1000000 THEN 'HIGH' WHEN total_net_amount >= 100000 THEN 'MEDIUM' ELSE 'LOW' END", "", "STRING", "10", "N", "2026-03-18 v1.0.0 初始创建", "渠道价值等级：高/中/低"),
        (20, "avg_inventory_turnover", "平均库存周转率", "CALCULATE", "", "", "", "", "", "CASE WHEN COALESCE(SUM(inv.avg_inventory_qty), 0) > 0 THEN SUM(total_quantity) / SUM(inv.avg_inventory_qty) ELSE 0 END", "", "DECIMAL", "10,2", "N", "2026-03-18 v1.0.0 初始创建", "平均库存周转率"),
        (21, "promotion_order_count", "促销订单数", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "order_id", "", "COUNT(DISTINCT CASE WHEN ord.promotion_id IS NOT NULL THEN ord.order_id END)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "促销订单数"),
        (22, "promotion_order_rate", "促销订单占比", "CALCULATE", "", "", "", "", "", "CASE WHEN total_orders > 0 THEN promotion_order_count * 100.0 / total_orders ELSE 0 END", "", "DECIMAL", "5,2", "N", "2026-03-18 v1.0.0 初始创建", "促销订单占比(%)"),
        (23, "first_time_customers", "首单客户数", "AGGREGATE", "dwd", "dwd_trade_orders_detail_day", "ord", "customer_id", "", "COUNT(DISTINCT CASE WHEN ord.is_first_order = 'Y' THEN ord.customer_id END)", "Y", "BIGINT", "", "N", "2026-03-18 v1.0.0 初始创建", "首单客户数"),
        (24, "repurchase_rate", "复购率", "CALCULATE", "", "", "", "", "", "CASE WHEN total_customers > 0 THEN (total_customers - first_time_customers) * 100.0 / total_customers ELSE 0 END", "", "DECIMAL", "5,2", "N", "2026-03-18 v1.0.0 初始创建", "复购率(%)"),
        (25, "etl_load_time", "ETL加载时间", "SYSTEM", "", "", "", "", "", "CURRENT_TIMESTAMP", "", "TIMESTAMP", "", "N", "2026-03-18 v1.0.0 初始创建", "ETL加载时间"),
    ]
    
    target_table = "dws_ecommerce_channel_sales_analysis_day"
    
    # 写入数据
    for row_idx, attr in enumerate(attributes, 2):
        ws.cell(row=row_idx, column=1, value=attr[0])
        ws.cell(row=row_idx, column=2, value=target_table)
        ws.cell(row=row_idx, column=3, value=attr[1])
        ws.cell(row=row_idx, column=4, value=attr[2])
        ws.cell(row=row_idx, column=5, value=attr[3])
        ws.cell(row=row_idx, column=6, value=attr[4])
        ws.cell(row=row_idx, column=7, value=attr[5])
        ws.cell(row=row_idx, column=8, value=attr[6])
        ws.cell(row=row_idx, column=9, value=attr[7])
        ws.cell(row=row_idx, column=10, value=attr[8])
        ws.cell(row=row_idx, column=11, value=attr[9])
        ws.cell(row=row_idx, column=12, value=attr[10])
        ws.cell(row=row_idx, column=13, value="数据开发工程师")
        ws.cell(row=row_idx, column=14, value=attr[11])
        ws.cell(row=row_idx, column=15, value=attr[12])
        ws.cell(row=row_idx, column=16, value=attr[13])
        ws.cell(row=row_idx, column=17, value=attr[14])
        ws.cell(row=row_idx, column=18, value=attr[15])
        
        # 设置行高和对齐
        ws.row_dimensions[row_idx].height = 40
        for col in range(1, 19):
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    
    # 设置表头行高
    ws.row_dimensions[1].height = 35
    
    # 保存
    wb.save("属性级Mapping_filled.xlsx")
    print("[OK] 属性级Mapping_filled.xlsx 已创建")

if __name__ == "__main__":
    print("开始生成填充好的Mapping Excel文档...")
    print()
    
    create_entity_mapping_excel()
    create_attribute_mapping_excel()
    
    print()
    print("="*60)
    print("所有文档生成完成!")
    print("="*60)
    print()
    print("生成的文件：")
    print("  1. 实体级Mapping_filled.xlsx - 填充好的实体级Mapping文档")
    print("  2. 属性级Mapping_filled.xlsx - 填充好的属性级Mapping文档（25个字段）")
    print()
    print("场景说明：")
    print("  - 业务场景：电商全渠道销售分析日汇总")
    print("  - 目标表：dws_ecommerce_channel_sales_analysis_day")
    print("  - 来源表：12个（5个事实表 + 7个维度表）")
    print("  - 加工逻辑：多表JOIN关联、聚合计算、窗口函数去重、")
    print("             CASE WHEN条件判断、公式计算、数据清洗")
    print()
